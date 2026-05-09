"""Excel report handling."""

import io
from pathlib import Path

import httpx
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from config import config

from .polling import get_api_answer
from .types import HWItem


async def create_report(token: str) -> io.BytesIO:
    """Create report containing all homeworks info.

    Args:
        token (str): user access token.

    Returns:
        Path: path to created report file.
    """
    async with httpx.AsyncClient(timeout=config.REQUEST_TIMEOUT) as session:
        homeworks = await get_api_answer(session, token, from_timestamp=0)
    data = pd.DataFrame(homeworks, columns=pd.Index(HWItem.__annotations__.keys()))
    Path("./temp.xlsx")
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer) as writer:
        data.to_excel(writer, sheet_name="Homeworks", index=False)
        worksheet: Worksheet = writer.sheets["Homeworks"]
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName="HomeworksTable", ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        worksheet.add_table(table)
    return buffer
